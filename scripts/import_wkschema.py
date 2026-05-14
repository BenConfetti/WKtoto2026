from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "wkschema.xlsx"
MATCHES_PATH = ROOT / "data" / "matches.json"
COMPETITION_PATH = ROOT / "data" / "competition.json"
COMPETITION_TIMEZONE = ZoneInfo("Europe/Amsterdam")


STAGE_LABELS = {
    "Round of 32": "Tweede ronde",
    "Round of 16": "Derde ronde",
    "Quarter final": "Kwartfinale",
    "Semi-Final": "Halve finale",
    "Third place": "Troostfinale",
    "Final": "Finale",
}

TEAM_TRANSLATIONS = {
    "Algeria": "Algerije",
    "Argentina": "Argentinië",
    "Australia": "Australië",
    "Austria": "Oostenrijk",
    "Belgium": "België",
    "Bosnia/Herzeg.": "Bosnië en Herzegovina",
    "Brazil": "Brazilië",
    "Canada": "Canada",
    "Cape Verde": "Kaapverdie",
    "Colombia": "Colombia",
    "Croatia": "Kroatie",
    "Curaçao": "Curaçao",
    "Czech Rep.": "Tsjechië",
    "DR Congo": "DR Congo",
    "Ecuador": "Ecuador",
    "Egypt": "Egypte",
    "England": "Engeland",
    "France": "Frankrijk",
    "Germany": "Duitsland",
    "Ghana": "Ghana",
    "Haiti": "Haïti",
    "IR Iran": "Iran",
    "Iraq": "Irak",
    "Ivory Coast": "Ivoorkust",
    "Japan": "Japan",
    "Jordan": "Jordanië",
    "Mexico": "Mexico",
    "Morocco": "Marokko",
    "Netherlands": "Nederland",
    "New Zealand": "Nieuw-Zeeland",
    "Norway": "Noorwegen",
    "Panama": "Panama",
    "Paraguay": "Paraguay",
    "Portugal": "Portugal",
    "Qatar": "Qatar",
    "Rep. of Korea": "Zuid-Korea",
    "Saudi Arabia": "Saoedi-Arabië",
    "Scotland": "Schotland",
    "Senegal": "Senegal",
    "South Africa": "Zuid-Afrika",
    "Spain": "Spanje",
    "Sweden": "Zweden",
    "Switzerland": "Zwitserland",
    "Tunisia": "Tunesië",
    "Turkey": "Turkije",
    "USA": "Verenigde Staten",
    "Uruguay": "Uruguay",
    "Uzbekistan": "Oezbekistan",
}


def parse_datetime(value) -> datetime | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return None
        dt = datetime.fromisoformat(text)

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=COMPETITION_TIMEZONE)

    return dt


def isoformat_utc(value: datetime) -> str:
    return value.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["DailySchedule"]

    matches = []
    current_stage = "Groepsfase"

    for row in sheet.iter_rows(min_row=1, values_only=True):
        date_value = row[1] if len(row) > 1 else None
        team_1 = row[3] if len(row) > 3 else None
        team_2 = row[4] if len(row) > 4 else None
        match_no = row[5] if len(row) > 5 else None
        team_code_1 = row[6] if len(row) > 6 else None
        team_code_2 = row[7] if len(row) > 7 else None
        venue = row[8] if len(row) > 8 else None

        if isinstance(date_value, str) and date_value.strip() in STAGE_LABELS and match_no is None:
            current_stage = STAGE_LABELS[date_value.strip()]
            continue

        if not isinstance(match_no, int):
            continue

        kickoff = parse_datetime(date_value)
        if kickoff is None:
            continue

        home_team = TEAM_TRANSLATIONS.get(team_1, team_1) if team_1 else team_code_1 or "Onbekend"
        away_team = TEAM_TRANSLATIONS.get(team_2, team_2) if team_2 else team_code_2 or "Onbekend"

        matches.append(
            {
                "id": f"match-{match_no:03d}",
                "matchNumber": match_no,
                "stage": current_stage,
                "sequence": 0,
                "homeTeam": home_team,
                "awayTeam": away_team,
                "city": venue or "",
                "kickoffAt": isoformat_utc(kickoff),
                "status": "scheduled",
                "homeScore": None,
                "awayScore": None,
                "homeSlotCode": team_code_1,
                "awaySlotCode": team_code_2,
            }
        )

    matches.sort(key=lambda match: (match["kickoffAt"], match["matchNumber"]))
    for index, match in enumerate(matches, start=1):
        match["sequence"] = index

    MATCHES_PATH.write_text(json.dumps(matches, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    competition = json.loads(COMPETITION_PATH.read_text(encoding="utf-8"))
    if matches:
        competition["predictionDeadline"] = matches[0]["kickoffAt"]
    COMPETITION_PATH.write_text(json.dumps(competition, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(json.dumps({"matchesImported": len(matches), "deadline": competition["predictionDeadline"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
